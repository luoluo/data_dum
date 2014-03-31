#encoding: utf-8
import sys
from sys import exc_info
import logging
from logger import Logger
from openpyxl import load_workbook
from openpyxl import cell
from openpyxl.cell import get_column_letter
class report_reader:
    def __init__(self, logger):
        self.logger= logger
        self.user_info_accurate_map = {}
        self.user_info_coverage_map = {}
        self.user_info_time_delay_map = {}
        self.wb = None
#   0           1         2         3              4            5        6             7                8               9         10          11              12           13         14
#time_period userid   username   website     register_time  audit_type status        audit_time      sent_time     is_illegal  illegal_type is_very_danger  register_time  is_new  l1_account
#time_period username l2_account l1_account  is_very_danger userid     register_time is_new          is_prism_found     9          10            11           12
#time_period username l1_account sources     is_very_danger userid     register_time is_new_register first_buy      sent_l2_audit l2_audit     audit_time  sent_time

    def dump_data(self, input_file):
        try:
            self.wb = load_workbook(input_file)
        except :
            e = exc_info()[1]
            self.logger.error("File not found: %s %s" % (input_file, e))
            return
        self.logger.info("data load start")
        self.read_accurate(self.user_info_accurate_map)
        self.read_coverage(self.user_info_coverage_map)
        self.read_time_delay(self.user_info_time_delay_map)
        self.logger.info("data lod end")

    def read_accurate(self,  user_info_accurate_map):
        cnt = 1
        ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
        self.logger.debug("reading accurate")
        for row in range(2, ws.get_highest_row()):
            items = []
            for col in range(1, ws.get_highest_column()+1):
                col = get_column_letter(col)
                items.append(ws['%s%s' % (col, row)].value)
#        #    is_very_danger
#            if items[6] == u'是':
#                items[6] = 1
#            elif row != 1:
#                items[6] = 0
#
#        #    is_audit_by_prism
#            if items[7] == u'是':
#                items[7] = 1
#            elif row != 1:
#                items[7] = 0

            if cnt == 10:
                break
            user_info_accurate_map[items[1]] = items
            self.logger.debug(items)
            cnt += 1
        self.logger.debug("read accurate end")
        print len(user_info_accurate_map.keys())
        for key in user_info_accurate_map.keys():
            print len(user_info_accurate_map[key])

    def read_coverage(self,  user_info_accurate_map):
        cnt = 1
        ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[1])
        user_info_coverage_map = {}
        self.logger.debug("reading coverage")
        for row in range(2,ws.get_highest_row()):
            items = []
            for col in range(1, ws.get_highest_column()+1):
                col = get_column_letter(col)
                items.append(ws['%s%s' % (col, row)].value)
            if cnt == 10:
                break
            user_info_coverage_map[items[4]] = items
            self.logger.debug(items)
            cnt += 1

        print len(user_info_coverage_map.keys())
        for key in user_info_coverage_map.keys():
            print len(user_info_coverage_map[key])
        self.logger.debug("read coverage end")

    def read_time_delay(self, user_info_time_delay_map):
        self.logger.debug("reading time delay")
        cnt = 1
        ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[2])
        for row in range(2, ws.get_highest_row()):
            items = []
            for col in range(1, ws.get_highest_column()+1):
                col = get_column_letter(col)
                items.append(ws['%s%s' % (col, row)].value)
            if cnt == 10:
                break
            user_info_time_delay_map[items[5]] = items
            self.logger.debug(items)
            cnt += 1
        self.logger.debug("read time delay end")


