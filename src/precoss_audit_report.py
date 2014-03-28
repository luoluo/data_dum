#encoding: utf-8
import sys
from openpyxl import load_workbook
from openpyxl import cell
from openpyxl.cell import get_column_letter
wb = load_workbook('ac.xlsx')
#print "worksheets ranges:", wb.get_named_ranges()
#print "worksheets names:", wb.get_sheet_names()

ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
#print "sheet title: ", ws.title
#print "sheet rowss: ", ws.get_highest_row()
#print "sheet columnss: ", ws.get_highest_column()
for row in range(1,ws.get_highest_row()):
	items = []
	for col in range(1, ws.get_highest_column()+1):
		col = get_column_letter(col)
		items.append(ws['%s%s' % (col, row)].value)
#		sys.stdout.write(ws['%s%s' % (col, row)].value)
#	sys.stdout.write("\t")
	if items[7] == u'是':
		print 'yes'
#	print ws.cell(2, 'A').vaule
